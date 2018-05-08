import pandas as pd
import natsort as nats
import matplotlib.pyplot as plt
import openpyxl
import numpy as np
#Read Excel file.
original_file = pd.read_excel(input('Please enter the name of your file, including the path to the file (i.e. C:/Desktop/Research/filename.xlsx):  '), header = None)
#Establish an empty dictionary which will be populated with well names and absorbance values.
valDict = {}
#Create a list containing the number of timepoints for a given kinetic read.  In the future, it would be optimal to adjust this to take input in case reads have different times.
timepoints = [n for n in range(int(input('How many timepoints do you have?  '))*2) if n%2 == 0]

#Iterate through each row of the DataFrame
for row in original_file.index:
	#Set the letter for the column name equal to the letter of the plate row at a given position
	letter = original_file.loc[row,0]
	#Don't include values when not a single letter, i.e. '<>' or 'Raw Data'
	if len(letter) > 1:
		continue
	#Iterate through each column in the DataFrame
	for column in original_file:
		#Skip the first column which contains the letters of the plate layout.
		if column == 0:
			continue
		#Create a key variable to be used in a dictionary.  Set this variable to the letter of the row + the number of the well (i.e. A + 5 = A5)	
		key = letter + str(column)
		#If they key is not already in the dictionary, create a dictionary key:value, which s empty at first.
		if key not in valDict:
			valDict[key] = []
		#Add the key (i.e. A5) and its associated values (data values at every timepoint)	
		valDict[key].append(original_file.loc[row,column])

#Make a new DataFrame which consists of the dictionary I made.  
reshaped_file = pd.DataFrame(valDict)
#Set the index of the new DataFrame to be the timepoints for a kinetic read.
reshaped_file.index = timepoints
#Rename the index column to have meaning.
reshaped_file.index.name = 'Time (minutes)'

#Until this point, the DataFrame columns aren't correctly ordered.  The reason is because the column labels are all strings.
#Because of this, A10 comes before A2 (because 1 < 2 and the next digit is ignored).  So we need to mdoify the names of the columns.
#Start by creating a list of the columns.
cols = reshaped_file.columns.tolist()
#Use the natsorted function to "naturally" sort the column names.  This is a 3rd party package that needs to be installed, and then imported.
#Now, A10 will come after A9.
new_cols = nats.natsorted(cols)
#Remake the DataFrame, using the naturally sorted columns.
reshaped_file = reshaped_file[new_cols]
reshaped_file = reshaped_file.dropna(how = 'all', axis = 1)
#Replace invalid values with NaN so copmutations can be performed.
reshaped_file = reshaped_file.replace('<Min', np.nan)
reshaped_file = reshaped_file.replace('>Max', np.nan)
#Write the new DataFrame to file.
#reshaped_file.to_excel('E:/OneDrive/Research/Rockwell_Lab/Python/041918_WTKO_Balbc_LDH_all_organized_RF.xlsx')
fancy_file_name = input('Please enter the name for your fancy pants organized file, including path and extension (i.e. C:/Desktop/Research/fancy_pants.xlsx):  ')
reshaped_file.to_excel(fancy_file_name)

#Load the organized workbook into 'wb' which will be used to add graphs to new tabs.
wb = openpyxl.load_workbook(fancy_file_name)

for column in reshaped_file:
	#print(column)
	#Make a list of x values for my plot equivalent to the timepoints of the curve.
    x_vals = reshaped_file.index.values
	#Make y values for the plot, equal to the absorbance values stored in each column.
    y_vals = reshaped_file[column]
	#Calling the figure function makes each plot its own figure, instead of adding each new line to the same figure.
    plt.figure()
	#Plot absorbance vs time
    plt.plot(x_vals, y_vals)
	#Label the x axis
    plt.xlabel('Time (minutes)')
	#Label the y axis
    plt.ylabel('Abs @ 450 nm')
	#Save the graphs
    plt.savefig(column + ".png", dpi = 100)
	
	#Create a new sheet in the Excel workbook.  This will create a new tab for each column.
    ws = wb.create_sheet(column)
	#Make an image object to be planted into Excel.
    img = openpyxl.drawing.image.Image(column + '.png')
	#Specify where the image will go in the tab, in this case, cell A1.
    img.anchor(ws['A1'])
	#Add the image.
    ws.add_image(img)
#Resave the organized workbook with the added images.	
wb.save(fancy_file_name)