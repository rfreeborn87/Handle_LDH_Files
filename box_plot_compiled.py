import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import seaborn as sns
import numpy as np
import openpyxl
#Import DataFrame from Excel file.
file_name = input('Please enter the file name, including the path, for which you\'d like to make a box plot:  ')

if file_name.endswith('.xlsx'):
    extensionless_file_name = file_name[:-5]
elif file_name.endswith('.xls'):
    extensionless_file_name = file_name[:-4]


compiled_data = pd.read_excel(file_name, sheet_name = "For_box")
#print(compiled_data.head())
#print(compiled_data['Genotype'].unique())
wt_indices = compiled_data['Genotype'] == 'WT'
WT = compiled_data.loc[wt_indices, :]
ko_indices = compiled_data['Genotype'] == 'KO'
KO = compiled_data.loc[ko_indices, :]

#pivoted = compiled_data.pivot(columns = 'Genotype', values = 'LDH_Activity')


#plt.boxplot(x = [WT['LDH_Activity'],KO['LDH_Activity']],labels = compiled_data['Genotype'].unique())#,y = 'LDH_Activity', kind = 'box')
#plt.ylabel('LDH Activity (mU/mL)')   
#plt.show()    

wb = openpyxl.load_workbook(file_name)
sns.set(style="whitegrid", color_codes=True)
ax = sns.swarmplot(x='Genotype', y='LDH_Activity', data=compiled_data)
ax = sns.boxplot(x='Genotype', y='LDH_Activity', data=compiled_data,
        showcaps=False,boxprops={'facecolor':'None'},
        showfliers=False,whiskerprops={'linewidth':0})
#plt.show()
#ax.set(yscale = 'log')
#ax.yaxis.set_major_locator(ticker.MultipleLocator(100))

#ax.set_yticks([10,50,100,200,300,400,500,600])
plt.savefig(extensionless_file_name + '.png', dpi = 100)
ws = wb.create_sheet('Figures')
img = openpyxl.drawing.image.Image(extensionless_file_name + '.png')
img.anchor(ws['A1'])
ws.add_image(img)



def remove_outliers():
    wt_q75, wt_q25 = np.percentile(WT['LDH_Activity'], [75 ,25])
    wt_iqr = wt_q75 - wt_q25
    
    wt_min = wt_q25 - (wt_iqr*1.5)
    wt_max = wt_q75 + (wt_iqr*1.5)

    WT['Outlier'] = 0
 
    WT.loc[WT['LDH_Activity'] < wt_min, 'Outlier'] = 1
    WT.loc[WT['LDH_Activity'] > wt_max, 'Outlier'] = 1

    WT_edited = WT[WT.Outlier != 1]
    #print(WT_edited.head())    

    ko_q75, ko_q25 = np.percentile(KO['LDH_Activity'], [75 ,25])
    ko_iqr = ko_q75 - ko_q25
 
    ko_min = ko_q25 - (ko_iqr*1.5)
    ko_max = ko_q75 + (ko_iqr*1.5)

    KO['Outlier'] = 0
 
    KO.loc[KO['LDH_Activity'] < ko_min, 'Outlier'] = 1
    KO.loc[KO['LDH_Activity'] > ko_max, 'Outlier'] = 1

    KO_edited = KO[KO.Outlier != 1]

    concatenated = pd.concat([WT_edited,KO_edited])
#    print(concatenated)
    plt.figure()
    sns.set(style="whitegrid", color_codes=True)
    new_ax = sns.swarmplot(x='Genotype', y='LDH_Activity', data=concatenated)
    new_ax = sns.boxplot(x='Genotype', y='LDH_Activity', data=concatenated,
            showcaps=False,boxprops={'facecolor':'None'},
            showfliers=False,whiskerprops={'linewidth':0})
    plt.savefig(extensionless_file_name + '_no_outliers.png', dpi = 100)
    img2 = openpyxl.drawing.image.Image(extensionless_file_name + '_no_outliers.png')
    img2.anchor(ws['K1'])
    ws.add_image(img2)
    #print(KO_edited.head())
if input("Do you also want a graph with the outliers removed? (y/n):  ").lower() == 'y':
    remove_outliers()

#elif input("Do you also want a graph with the outliers removed? (y/n):  ").lower() == 'n':
#    break
#else:
#    print("That was a y or n question.  We don't accept full words or wrong letters here.")

wb.save(file_name)
